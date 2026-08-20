"""Excelガイドのテーブル定義をWindows Excel互換に保つ回帰テスト。"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pytest
from PIL import Image
from openpyxl import Workbook

from generate_evaluation_telemetry_guide import (
    _add_table,
    _action_history,
    _build_workbook,
    _select_default_inputs,
    _validate_workbook_tables,
)


def _workbook_with_table() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(("項目", "説明"))
    worksheet.append(("STATUS", "終了状態"))
    _add_table(worksheet, name="TelemetryGuide", ref="A1:B2")
    return workbook


def test_default_inputs_prefer_complete_canonical_set_and_fall_back_as_a_set(
    tmp_path: Path,
) -> None:
    filenames = ("frame.png", "steps.jsonl", "evaluation.json")
    canonical = tuple(tmp_path / "canonical" / name for name in filenames)
    legacy = tuple(tmp_path / "legacy" / name for name in filenames)
    for path in legacy:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.touch()

    assert _select_default_inputs(canonical, legacy) == legacy

    canonical[0].parent.mkdir(parents=True, exist_ok=True)
    canonical[0].touch()
    assert _select_default_inputs(canonical, legacy) == legacy

    canonical[1].touch()
    canonical[2].touch()
    assert _select_default_inputs(canonical, legacy) == canonical


def test_table_rejects_duplicate_worksheet_auto_filter() -> None:
    workbook = _workbook_with_table()
    worksheet = workbook.active
    worksheet.auto_filter.ref = "A1:B2"

    with pytest.raises(ValueError, match="オートフィルターを重複定義"):
        _validate_workbook_tables(workbook)


def test_table_serializes_only_its_own_auto_filter() -> None:
    workbook = _workbook_with_table()
    _validate_workbook_tables(workbook)
    output = BytesIO()
    workbook.save(output)

    with ZipFile(output) as archive:
        worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
        table_xml = archive.read("xl/tables/table1.xml")
        workbook_xml = archive.read("xl/workbook.xml")

    assert b"<autoFilter" not in worksheet_xml
    assert table_xml.count(b"<autoFilter") == 1
    assert b"_xlnm._FilterDatabase" not in workbook_xml


def test_builds_single_combined_guide_sheet() -> None:
    workbook = _build_workbook(
        image=Image.new("RGB", (920, 600), "white"),
        example={},
    )

    assert workbook.sheetnames == ["テレメトリ説明"]
    worksheet = workbook["テレメトリ説明"]
    assert len(worksheet._images) == 1
    assert (worksheet._images[0].width, worksheet._images[0].height) == (920, 600)
    assert worksheet.auto_filter.ref is None
    assert set(worksheet.tables) == {
        "TelemetryFieldGuide",
        "StatusLegend",
        "ActionColorLegend",
    }

    field_table = worksheet.tables["TelemetryFieldGuide"]
    status_table = worksheet.tables["StatusLegend"]
    action_table = worksheet.tables["ActionColorLegend"]
    assert field_table.ref == "A41:D62"
    assert status_table.ref == "A68:E76"
    assert action_table.ref == "A80:E86"
    assert field_table.autoFilter is not None
    assert status_table.autoFilter is not None
    assert action_table.autoFilter is not None

    # Each table has one header row; the remaining rows are the requested
    # 21 field explanations, 8 statuses, and 6 action-history colors.
    assert worksheet.max_column >= 6
    assert [worksheet.cell(41, column).value for column in range(1, 5)] == [
        "表示ラベル",
        "JSONフィールド／出典",
        "frame_000030例",
        "説明",
    ]
    assert worksheet["A41"].value != "番号"
    assert worksheet["A62"].value == "T/B"
    assert worksheet["A68"].value == "表示"
    assert worksheet["A76"].value == "TERMINATED"
    assert worksheet["A80"].value == "帯"
    assert worksheet["A86"].value == "T/B"
    assert "GIFとMP4" in worksheet["A90"].value
    assert worksheet["C42"].value == "道路・車両・車線区画・車両履歴"
    assert "対向方向を分ける黄色連続中央線" in worksheet["D42"].value
    assert "ルート線ではありません" in worksheet["D42"].value
    assert "ルートの目安" not in worksheet["D42"].value
    # The item-label column is descriptive text, not a color legend.  Every
    # data cell therefore uses the same neutral font and white fill; the
    # STATUS/ACTION color tables below retain their own semantic colors.
    label_cells = [worksheet.cell(row, 1) for row in range(42, 63)]
    assert {cell.fill.fill_type for cell in label_cells} == {"solid"}
    assert {cell.fill.fgColor.type for cell in label_cells} == {"rgb"}
    assert {cell.fill.fgColor.rgb for cell in label_cells} == {"00FFFFFF"}
    assert {cell.font.color.type for cell in label_cells if cell.font.color} == {"rgb"}
    assert {cell.font.color.rgb for cell in label_cells if cell.font.color} == {"0017212B"}
    assert len({cell.style_id for cell in label_cells}) == 1
    assert worksheet["A69"].font.color.rgb == "0046CD73"
    assert worksheet["D69"].fill.fgColor.rgb == "0046CD73"
    assert worksheet["C82"].fill.fgColor.rgb == "00F5A541"
    assert all(worksheet.row_dimensions[row].height == 15 for row in range(5, 35))
    assert "A1:F1" in {str(cell_range) for cell_range in worksheet.merged_cells.ranges}
    assert set(worksheet.column_dimensions) == set("ABCDEF")
    assert worksheet.print_title_rows == "$1:$2"
    assert worksheet.print_area == "'テレメトリ説明'!$A$1:$F$92"

    cell_text = "\n".join(
        str(cell.value)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "GIF PLAYBACK" not in cell_text
    assert "ACTION DT" not in cell_text
    assert "action_duration_seconds" not in cell_text
    assert all(
        forbidden not in cell_text
        for forbidden in (
            "EVALUATION TELEMETRY",
            "SIMULATION TIMING (runtime config)",
            "VEHICLE / APPLIED ACTION",
            "CURRENT SEGMENT (runtime)",
            "TASK",
            "日本語の意味",
            "読み方・注意",
            "読み方",
        )
    )
    assert "同じIDを続けても増えません" in worksheet["D51"].value
    assert "道路舗装全体の幅と断定しません" in worksheet["D55"].value
    assert "本評価の派生表示" in worksheet["B69"].value
    assert "総合フラグ" in worksheet["C74"].value
    assert worksheet["C83"].fill.fgColor.rgb == "00737D8C"
    assert worksheet["C86"].fill.fgColor.rgb == "00737D8C"
    assert worksheet["D83"].value == "737D8C"
    assert worksheet["D86"].value == "737D8C"
    _validate_workbook_tables(workbook)


def test_action_history_uses_runtime_ceil_for_non_integer_window() -> None:
    config = {
        "discrete_action": True,
        "use_multi_discrete": False,
        "discrete_steering_dim": 3,
        "discrete_throttle_dim": 3,
    }
    rows = [
        {"episode": 1, "step": step, "action_id": step}
        for step in range(1, 7)
    ]
    history = _action_history(
        rows,
        {"episode": 1, "step": 6, "control_hz": 2.25},
        config,
    )

    assert len(history) == 5  # ceil(2 s × 2.25 Hz), not round(...)
    assert [item.action_id for item in history] == [2, 3, 4, 5, 6]
