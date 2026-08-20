"""評価テレメトリの読み方を編集可能なExcel資料として生成する。

このスクリプトは、既存評価のJSON/JSONLと記録済みフレームを入力にする。
フレームの左側（地図）だけを元画像から取り出し、現行の
``compose_telemetry_panel`` で右側パネルを再合成してから、Excelへ画像として
埋め込む。説明文や例の値はすべてセルへ書き込むため、Excel上で直接編集できる。

Usage::

    .venv/bin/python generate_evaluation_telemetry_guide.py
    .venv/bin/python generate_evaluation_telemetry_guide.py --output /tmp/guide.xlsx
"""

from __future__ import annotations

import argparse
from io import BytesIO
import json
import math
from pathlib import Path
import re
from typing import Any, Mapping, Sequence

from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

from evaluation_visualization import (
    ACTION_HISTORY_SECONDS,
    compose_telemetry_panel,
    decode_discrete_action,
)
from configs.phase0_config import OFFICIAL_EVALUATION_OUTPUT_DIR


PROJECT_ROOT = Path(__file__).resolve().parent
CANONICAL_FRAME_PATH = (
    OFFICIAL_EVALUATION_OUTPUT_DIR
    / "episodes/episode_0001_scenario_000005/frames/frame_000030.png"
)
CANONICAL_STEPS_PATH = OFFICIAL_EVALUATION_OUTPUT_DIR / "evaluation_steps.jsonl"
CANONICAL_EVALUATION_PATH = OFFICIAL_EVALUATION_OUTPUT_DIR / "evaluation.json"
LEGACY_FRAME_PATH = PROJECT_ROOT / "outputs/phase0_official_evaluation_frames/frame_000030.png"
LEGACY_STEPS_PATH = PROJECT_ROOT / "outputs/phase0_official_evaluation_steps.jsonl"
LEGACY_EVALUATION_PATH = PROJECT_ROOT / "outputs/phase0_official_evaluation.json"


def _select_default_inputs(
    canonical: tuple[Path, Path, Path],
    legacy: tuple[Path, Path, Path],
) -> tuple[Path, Path, Path]:
    """Use one complete input set, preferring the current output layout."""

    if all(path.is_file() for path in canonical):
        return canonical
    if all(path.is_file() for path in legacy):
        return legacy
    # Keep the current layout authoritative when neither set is complete, so
    # the eventual error points users to the path produced by a new evaluation.
    return canonical


DEFAULT_FRAME_PATH, DEFAULT_STEPS_PATH, DEFAULT_EVALUATION_PATH = (
    _select_default_inputs(
        (CANONICAL_FRAME_PATH, CANONICAL_STEPS_PATH, CANONICAL_EVALUATION_PATH),
        (LEGACY_FRAME_PATH, LEGACY_STEPS_PATH, LEGACY_EVALUATION_PATH),
    )
)
DEFAULT_OUTPUT_PATH = PROJECT_ROOT / "EVALUATION_TELEMETRY_GUIDE.xlsx"

MAP_WIDTH = 600
FRAME_NUMBER = 30
PANEL_WIDTH = 320


# The palette is intentionally simple so a user can edit it without needing to
# understand an Excel theme.  Values are ARGB strings, as expected by openpyxl.
NAVY = "17324D"
CYAN = "007E9E"
CYAN_SOFT = "E6F7FB"
ORANGE = "BD6500"
ORANGE_SOFT = "FFF1DF"
GREEN = "46CD73"
GREEN_SOFT = "EAF7EF"
RED = "EB5050"
RED_SOFT = "FCECEC"
GRAY = "778392"
# This is the exact neutral/zero color used by the runtime ACTION HISTORY
# panel.  Keep it separate from GRAY, which is a document-style color.
HISTORY_GRAY = "737D8C"
GRAY_SOFT = "F1F3F5"
BLUE = "3C96F5"
BLUE_SOFT = "EAF3FF"
WHITE = "FFFFFF"
BLACK = "17212B"
LINE = "D8E0E8"


_TABLE_NAME_RE = re.compile(r"^[A-Za-z_\\][A-Za-z0-9_.]*$")
_CELL_REFERENCE_RE = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*$")
_HEADER_BREAKS = ("\r", "\n", "\x0b", "\x0c", "\u2028", "\u2029")
_ILLEGAL_XML_CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="評価テレメトリ説明資料を編集可能なExcelへ出力します。"
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"出力先xlsx（既定: {DEFAULT_OUTPUT_PATH.name}）",
    )
    parser.add_argument(
        "--frame",
        type=Path,
        default=DEFAULT_FRAME_PATH,
        help=f"例に使うPNG（既定: {DEFAULT_FRAME_PATH}）",
    )
    parser.add_argument(
        "--steps",
        type=Path,
        default=DEFAULT_STEPS_PATH,
        help=f"ステップJSONL（既定: {DEFAULT_STEPS_PATH}）",
    )
    parser.add_argument(
        "--evaluation",
        type=Path,
        default=DEFAULT_EVALUATION_PATH,
        help=f"評価JSON（既定: {DEFAULT_EVALUATION_PATH}）",
    )
    return parser.parse_args()


def _resolve_input(path: Path) -> Path:
    """Resolve relative input paths relative to the project, then validate them."""

    resolved = path if path.is_absolute() else PROJECT_ROOT / path
    if not resolved.is_file():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {resolved}")
    return resolved


def _resolve_output(path: Path) -> Path:
    """Resolve a relative output path relative to the current project."""

    resolved = path if path.is_absolute() else PROJECT_ROOT / path
    resolved.parent.mkdir(parents=True, exist_ok=True)
    return resolved


def _read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as file_obj:
        payload = json.load(file_obj)
    if not isinstance(payload, dict):
        raise ValueError(f"評価JSONのトップレベルはobjectである必要があります: {path}")
    return payload


def _read_step_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as file_obj:
        for line_number, line in enumerate(file_obj, start=1):
            if not line.strip():
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"JSONL {path} の{line_number}行目を読めません") from exc
            if isinstance(row, dict):
                rows.append(row)
    if not rows:
        raise ValueError(f"ステップJSONLに行がありません: {path}")
    return rows


def _example_row(
    rows: Sequence[Mapping[str, Any]],
    *,
    episode: int = 1,
    step: int = FRAME_NUMBER,
) -> dict[str, Any]:
    """Return a copy of the requested row and normalize current timing metadata."""

    for row in rows:
        if int(row.get("episode", -1)) == episode and int(row.get("step", -1)) == step:
            example = dict(row)
            # The compatibility field is retained in JSON by evaluate.py, but
            # the current artifacts are deliberately synchronized at 1x speed.
            example["gif_playback_vs_simulation"] = 1.0
            return example
    raise ValueError(f"episode={episode}, step={step} のテレメトリ行がありません")


def _runtime_config(evaluation: Mapping[str, Any]) -> dict[str, Any]:
    config = dict(evaluation.get("environment_config") or {})
    if not config:
        raise ValueError("評価JSONにenvironment_configがありません")
    # The official output does not need these keys, but supplying explicit
    # defaults makes action decoding stable for older evaluation JSON files.
    config.setdefault("use_multi_discrete", False)
    config.setdefault("vehicle_config", {"enable_reverse": False})
    return config


def _action_history(
    rows: Sequence[Mapping[str, Any]],
    example: Mapping[str, Any],
    config: Mapping[str, Any],
) -> list[Any]:
    episode = int(example["episode"])
    step = int(example["step"])
    control_hz = float(example.get("control_hz") or 0.0)
    history_count = max(1, int(math.ceil(ACTION_HISTORY_SECONDS * control_hz)))
    history_rows = [
        row
        for row in rows
        if int(row.get("episode", -1)) == episode
        and 1 <= int(row.get("step", -1)) <= step
    ][-history_count:]
    return [
        decode_discrete_action(int(row["action_id"]), config)
        for row in history_rows
    ]


def _updated_frame(
    source_path: Path,
    example: Mapping[str, Any],
    action_history: Sequence[Any],
) -> PILImage.Image:
    """Rebuild the current 920x600 panel while preserving the source map."""

    with PILImage.open(source_path) as source:
        source_rgb = source.convert("RGB")
        if source_rgb.width < MAP_WIDTH or source_rgb.height <= 0:
            raise ValueError(f"フレームサイズが想定より小さいです: {source_rgb.size}")
        map_part = source_rgb.crop((0, 0, MAP_WIDTH, source_rgb.height))
        composite = compose_telemetry_panel(
            map_part,
            example,
            action_history,
            panel_width=PANEL_WIDTH,
        )
    return PILImage.fromarray(composite, mode="RGB")


def _thin(color: str = LINE) -> Side:
    return Side(style="thin", color=color)


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _style_title(cell: Any, *, size: int = 18, color: str = NAVY) -> None:
    cell.font = Font(name="Yu Gothic", size=size, bold=True, color=color)
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def _style_header_row(ws: Any, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Yu Gothic", size=11, bold=True, color=WHITE)
        cell.fill = _fill(CYAN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=_thin(CYAN), bottom=_thin(CYAN))


def _style_body(ws: Any, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        fill = _fill("F8FBFC" if (row - start_row) % 2 == 0 else WHITE)
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = Font(name="Yu Gothic", size=10, color=BLACK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=_thin())


def _set_print_layout(ws: Any, *, title_rows: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    if title_rows:
        ws.print_title_rows = title_rows


def _add_table(
    ws: Any,
    *,
    name: str,
    ref: str,
    style: str = "TableStyleMedium2",
) -> None:
    headers = _validate_table_definition(ws, name=name, ref=ref)
    table = Table(displayName=name, ref=ref)
    # Define the filter and columns through public openpyxl objects so that the
    # in-memory contract is the same one written into the OOXML package.
    table.autoFilter = AutoFilter(ref=ref)
    table.tableColumns = [
        TableColumn(id=index, name=header)
        for index, header in enumerate(headers, start=1)
    ]
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _validate_table_definition(ws: Any, *, name: str, ref: str) -> list[str]:
    """Validate the parts of an Excel table that Excel rejects on open.

    Excel table headers are serialized as XML attributes, so an empty,
    duplicated, non-string, or line-breaking header can make the table
    unreadable even though openpyxl can still load the package.  The table
    name and range are checked here as well, before anything is added to the
    worksheet.
    """

    if not isinstance(name, str) or not name:
        raise ValueError("Excelテーブル名は空でない文字列である必要があります")
    if len(name) > 255 or not _TABLE_NAME_RE.fullmatch(name):
        raise ValueError(f"Excelテーブル名が不正です: {name!r}")
    if name.upper() in {"R", "C"} or _CELL_REFERENCE_RE.fullmatch(name):
        raise ValueError(f"Excelテーブル名がセル参照と衝突しています: {name!r}")

    if not isinstance(ref, str) or not ref:
        raise ValueError(f"テーブル {name!r} のrefが空です")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    except ValueError as exc:
        raise ValueError(f"テーブル {name!r} のrefが不正です: {ref!r}") from exc
    if None in (min_col, min_row, max_col, max_row) or min_col < 1 or min_row < 1:
        raise ValueError(f"テーブル {name!r} のrefが不正です: {ref!r}")
    if max_col < min_col or max_row < min_row:
        raise ValueError(f"テーブル {name!r} のrefが矩形ではありません: {ref!r}")

    headers: list[str] = []
    seen: set[str] = set()
    for column in range(min_col, max_col + 1):
        value = ws.cell(row=min_row, column=column).value
        if not isinstance(value, str) or not value.strip():
            raise ValueError(
                f"テーブル {name!r} のヘッダーが空または文字列ではありません: "
                f"{get_column_letter(column)}{min_row}={value!r}"
            )
        if len(value) > 255 or any(marker in value for marker in _HEADER_BREAKS):
            raise ValueError(
                f"テーブル {name!r} のヘッダーに改行または長すぎる文字列があります: "
                f"{value!r}"
            )
        if _ILLEGAL_XML_CONTROL_RE.search(value):
            raise ValueError(
                f"テーブル {name!r} のヘッダーにXML非対応の制御文字があります: "
                f"{value!r}"
            )
        duplicate_key = value.casefold()
        if duplicate_key in seen:
            raise ValueError(f"テーブル {name!r} のヘッダーが重複しています: {value!r}")
        seen.add(duplicate_key)
        headers.append(value)
    return headers


def _validate_workbook_tables(workbook: Workbook) -> None:
    """Check every table after construction, including cross-sheet names."""

    seen_names: set[str] = set()
    for ws in workbook.worksheets:
        if ws.tables and ws.auto_filter.ref is not None:
            raise ValueError(
                f"シート {ws.title!r} ではExcelテーブルとワークシートの"
                "オートフィルターを重複定義できません"
            )
        for table in ws.tables.values():
            name = table.displayName or table.name
            normalized_name = name.casefold()
            if normalized_name in seen_names:
                raise ValueError(f"Excelテーブル名が重複しています: {name!r}")
            seen_names.add(normalized_name)
            headers = _validate_table_definition(ws, name=name, ref=table.ref)
            if len(table.tableColumns) != len(headers):
                raise ValueError(
                    f"テーブル {name!r} のtableColumns countが不一致です: "
                    f"headers={len(headers)}, columns={len(table.tableColumns)}"
                )
            actual_names = [column.name for column in table.tableColumns]
            if actual_names != headers:
                raise ValueError(
                    f"テーブル {name!r} の列名がヘッダーと一致しません: "
                    f"{actual_names!r} != {headers!r}"
                )


def _set_widths(ws: Any, widths: Mapping[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _write_merged_text(ws: Any, cell_range: str, value: str, *, fill: str | None = None) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":", 1)[0]]
    cell.value = value
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.font = Font(name="Yu Gothic", size=10, color=BLACK)
    if fill:
        cell.fill = _fill(fill)


def _panel_legend_rows(example: Mapping[str, Any]) -> list[tuple[Any, ...]]:
    """Explain data-bearing labels; omit section headings and the redundant duration row."""

    return [
        (
            "地図（左側）",
            "TopDownRenderer / map renderer",
            "道路・車両・車線区画・車両履歴",
            "TopDownRendererが上から描く地図領域です。緑の矩形は対象車、薄い緑はTopDownRendererが重ねた過去位置、濃灰はこのPGマップのSIDE/GUARDRAIL外側線、薄灰は主に車線区切りです。オレンジはこの画像では対向方向を分ける黄色連続中央線で、ルート線ではありません。地図・描画設定により見え方は変わります。",
        ),
        (
            "POST-STEP",
            "step / horizon / sim_time_seconds",
            f"POST-STEP {example.get('step', 'N/A')}/{example.get('horizon', 'N/A')}  t={float(example.get('sim_time_seconds', 0.0)):.2f}s",
            "env.step()でアクションを適用した後の、1から始まるstep番号とシミュレーション時刻tです。k番目はk番目の制御区間の終了状態で、tは実時間（wall clock）ではありません。",
        ),
        (
            "ACTION APPLIED",
            "interval_start_seconds / sim_time_seconds",
            f"[{float(example.get('interval_start_seconds', 0.0)):.2f}, {float(example.get('sim_time_seconds', 0.0)):.2f}]s",
            "このフレームの状態へ至るまでに適用した1制御区間[ t-dt, t ]です。終端はPOST-STEPのtと一致し、dtは1制御周期です。",
        ),
        (
            "PHYSICS",
            "physics_hz",
            f"{float(example.get('physics_hz', 0.0)):.1f} Hz",
            "物理エンジンの小刻みな更新（substep）周波数です。値は1 / physics_world_step_size [Hz]で、1物理更新の時間はその逆数です。",
        ),
        (
            "CONTROL",
            "control_hz",
            f"{float(example.get('control_hz', 0.0)):.1f} Hz",
            "env.step()およびポリシーがアクションを決める周波数です。1 / (物理step × decision_repeat)で導かれ、独立した別設定ではありません。",
        ),
        (
            "SPEED",
            "speed_km_h / speed_m_s",
            f"{float(example.get('speed_km_h', 0.0)):.1f} km/h ({float(example.get('speed_m_s', 0.0)):.2f} m/s)",
            "同梱MetaDrive 0.4.3派生ソースのinfo[\"velocity\"]はvehicle.speed（m/s）です。本評価は同じ値を×3.6してkm/hも表示します。車両の速度の大きさで、進行方向の符号は持ちません。",
        ),
        (
            "ACTION",
            "action_id / action_label",
            f"{example.get('action_id', 'N/A')}  {example.get('action_label', 'N/A')}",
            "離散アクションIDと、操舵・縦方向操作へ復号したラベルです。IDの対応はactionの次元設定に依存し、現行の3×3設定ではID 7がSTRAIGHT + THROTTLEです。",
        ),
        (
            "STEERING",
            "applied_steering",
            f"{float(example.get('applied_steering', 0.0)):+.2f}",
            "環境へ適用された正規化操舵値[-1, 1]です。本評価の規約では+がLEFT、-がRIGHTで、角度（度）ではありません。",
        ),
        (
            "THROTTLE/BRK",
            "applied_throttle_brake",
            f"{float(example.get('applied_throttle_brake', 0.0)):+.2f}",
            "正規化された縦方向の操作値で、m/s²の加速度ではありません。enable_reverse=Falseでは+がengine/throttle、-がbrakeです。0入力でもBaseVehicleは各輪に小さい固定brake 2.0を設定するため、完全な無作動とは限りません。reverse有効時は負値が後退力になり得ます。",
        ),
        (
            "SWITCH COUNT",
            "action_switch_count",
            f"{example.get('action_switch_count', 'N/A')} (cumulative)",
            "reset後に、直前の離散action IDから別のIDへ変化した回数の累積です。最初の制御判断は数えず、同じIDを続けても増えません。",
        ),
        (
            "SWITCH RATE",
            "action_switches_per_second",
            f"{float(example.get('action_switches_per_second', 0.0)):.2f}/s (sim avg)",
            "SWITCH COUNTをシミュレーション経過時間tで割ったエピソード平均です。瞬時の切替率でも実時間（wall clock）基準でもありません。",
        ),
        (
            "LANE WIDTH",
            "lane_width_m",
            f"{float(example.get('lane_width_m', 0.0)):.2f} m",
            "現在のナビゲーション（navigation）が返す、1本の車線の幅[m]です。道路・マップ設定に依存します。",
        ),
        (
            "LANES",
            "lane_count_one_way",
            f"{example.get('lane_count_one_way', 'N/A')} one-way",
            "ナビゲーション（navigation）のget_current_lane_num()が返す、現在進行方向の参照車線数です。対向方向の車線は含まず、ナビゲーションモジュールの実装に依存します。",
        ),
        (
            "DRIVABLE",
            "current_segment_drivable_width_m",
            f"{float(example.get('current_segment_drivable_width_m', 0.0)):.2f} m",
            "ナビゲーション（navigation）のget_current_lateral_range()が返す、現在の参照範囲の横幅[m]です。通常のPG区間ではlane width×参照車線数、合流・分岐（merge/split）等ではレイキャスト（raycast）由来の場合があります。取得できないときの本評価の代替値も同じ積です。道路舗装全体の幅と断定しません。",
        ),
        (
            "CENTER->EDGE",
            "center_to_left_boundary_m / center_to_right_boundary_m",
            f"L {float(example.get('center_to_left_boundary_m', 0.0)):.2f} m  R {float(example.get('center_to_right_boundary_m', 0.0)):.2f} m",
            "車両中心からナビゲーション（navigation）の横方向参照範囲の左端・右端までの符号付き距離[m]です。車幅は引いていません。通常はL+R=DRIVABLEとなり、範囲外では負値になり得ます。",
        ),
        (
            "ROUTE",
            "route_completion",
            f"{float(example.get('route_completion', 0.0)) * 100:.1f}%",
            "MetaDriveが設定したルート上の進捗です。NodeNetworkNavigationでは、最短経路上のtravelled length ÷ total lengthを%表示します。直線距離や残距離ではなく、特殊な移動では単純な0–100%に収まらない場合があります。",
        ),
        (
            "REWARD",
            "step_reward / cumulative_reward",
            f"{float(example.get('step_reward', 0.0)):+.3f}  total {float(example.get('cumulative_reward', 0.0)):+.3f}",
            "左はenv.step()が返した今回のreward、totalは本評価スクリプトがその戻り値をreset後から独自に加算した累積値です。表示にはinfo[\"episode_reward\"]を使っていません。rewardの意味・尺度は環境configに依存します。",
        ),
        (
            "STATUS",
            "status / termination flags",
            str(example.get("status", "N/A")),
            "元フラグそのものではなく、本評価が表示用に作る派生表示名です。優先順はSUCCESS > OUT_OF_ROAD > CRASH_VEHICLE > CRASH_OBJECT > CRASH > MAX_STEP > TERMINATED > RUNNINGです。",
        ),
        (
            "ACTION HISTORY (last 2s)",
            "action_history",
            "直近2シミュレーション秒",
            "元のaction IDの履歴です。最大ceil(2×CONTROL)セル（現行10Hzなら20）で、左が古く右が新しく、現在stepを含みます。エピソード開始直後は履歴が短くなります。上段Sが操舵、下段T/Bが縦方向操作です。",
        ),
        (
            "S",
            "action history upper row",
            "blue / orange / gray",
            "本評価独自の上段色です。青はpositive/LEFT、オレンジはnegative/RIGHT、grayはzeroです。色はMetaDrive公式の一般規約ではなく、このパネルの表示規約です。",
        ),
        (
            "T/B",
            "action history lower row",
            "green / red / gray",
            "本評価独自の下段色です。緑はpositive/THROTTLE、赤はnegative（現行enable_reverse=FalseではBRAKE）、grayはzeroです。reverse設定では負値の意味が後退力に変わり得ます。",
        ),
    ]


def _write_combined_guide(ws: Any, image: PILImage.Image, example: Mapping[str, Any]) -> None:
    """Write the compact, single-sheet guide requested by the user.

    The image is kept at the top at its native 920x600 size.  The editable
    field table follows it, then the STATUS and ACTION HISTORY color tables,
    and finally the media/simulation timing note.
    """

    _set_print_layout(ws, title_rows="1:2")
    ws.sheet_view.zoomScale = 75
    _set_widths(
        ws,
        {
            "A": 31,
            "B": 34,
            "C": 31,
            "D": 70,
            "E": 30,
            "F": 18,
        },
    )
    ws.merge_cells("A1:F1")
    ws["A1"] = "MetaDrive 評価テレメトリ表示ガイド（編集可能版）"
    _style_title(ws["A1"], size=18, color=NAVY)
    ws["A1"].fill = _fill(CYAN_SOFT)
    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "上の画像は現行のデバッグ表示を再合成したものです。"
        "画像の下に、右パネルの項目説明、STATUS・色、メディア時間の説明をまとめています。"
        "説明文・例の値はセル上で直接編集できます。"
    )
    ws["A2"].font = Font(name="Yu Gothic", size=11, color=BLACK)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34

    ws.merge_cells("A4:F4")
    ws["A4"] = "現行の評価テレメトリ表示（frame_000030例）"
    ws["A4"].font = Font(name="Yu Gothic", size=13, bold=True, color=CYAN)
    ws["A4"].fill = _fill(CYAN_SOFT)
    ws["A4"].alignment = Alignment(vertical="center")

    image_buffer = BytesIO()
    image.save(image_buffer, format="PNG")
    image_buffer.seek(0)
    xl_image = XLImage(image_buffer)
    xl_image.width = 920
    xl_image.height = 600
    ws.add_image(xl_image, "A5")
    for row in range(5, 35):
        ws.row_dimensions[row].height = 15

    ws.merge_cells("A35:F35")
    ws["A35"] = (
        "画像は現行の右パネルを含む920×600のフレームです。"
        "1制御ステップ=1フレームとして、動画の累積再生時間と表示tを一致させます。"
    )
    ws["A35"].font = Font(name="Yu Gothic", size=10, italic=True, color=GRAY)
    ws["A35"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[35].height = 32

    # ------------------------------------------------------------------
    # 項目説明: visible data fields only (section-heading rows are omitted).
    # ------------------------------------------------------------------
    field_title_row = 38
    ws.merge_cells(start_row=field_title_row, start_column=1, end_row=field_title_row, end_column=6)
    ws.cell(field_title_row, 1).value = "項目説明（右パネルを上から読む）"
    _style_title(ws.cell(field_title_row, 1), size=16, color=NAVY)
    ws.cell(field_title_row, 1).fill = _fill(CYAN_SOFT)
    ws.merge_cells(start_row=field_title_row + 1, start_column=1, end_row=field_title_row + 1, end_column=6)
    ws.cell(field_title_row + 1, 1).value = (
        "表示ラベルを上から順に並べています。出典フィールド、例の表示、説明を確認してください。"
        "フィルターで領域ごとに絞り込めます。"
    )
    ws.cell(field_title_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(field_title_row + 1, 1).font = Font(name="Yu Gothic", size=10, color=BLACK)
    ws.row_dimensions[field_title_row + 1].height = 32

    field_header_row = field_title_row + 3
    field_headers = (
        "表示ラベル",
        "JSONフィールド／出典",
        "frame_000030例",
        "説明",
    )
    for column, value in enumerate(field_headers, start=1):
        ws.cell(field_header_row, column).value = value
    _style_header_row(ws, field_header_row, 1, len(field_headers))
    field_rows = _panel_legend_rows(example)
    for row_index, values in enumerate(field_rows, start=field_header_row + 1):
        for column, value in enumerate(values, start=1):
            ws.cell(row_index, column).value = value
        ws.row_dimensions[row_index].height = 72
    field_data_start = field_header_row + 1
    field_data_end = field_header_row + len(field_rows)
    _style_body(ws, field_data_start, field_data_end, 1, len(field_headers))
    # The label column is a lookup key, not a status legend.  Keep every
    # data cell visually neutral so its occasional section-like colors do not
    # imply a meaning that is not part of the telemetry definition.
    for row_index in range(field_data_start, field_data_end + 1):
        label_cell = ws.cell(row_index, 1)
        label_cell.fill = _fill(WHITE)
        label_cell.font = Font(name="Yu Gothic", size=10, color=BLACK)
        label_cell.alignment = Alignment(vertical="top", wrap_text=True)
    _add_table(
        ws,
        name="TelemetryFieldGuide",
        ref=f"A{field_header_row}:D{field_data_end}",
    )

    # ------------------------------------------------------------------
    # STATUS and ACTION HISTORY color legends.
    # ------------------------------------------------------------------
    status_title_row = field_data_end + 3
    ws.merge_cells(start_row=status_title_row, start_column=1, end_row=status_title_row, end_column=6)
    ws.cell(status_title_row, 1).value = "STATUS・色"
    ws.cell(status_title_row, 1).font = Font(name="Yu Gothic", size=16, bold=True, color=NAVY)
    ws.cell(status_title_row, 1).fill = _fill(ORANGE_SOFT)
    ws.merge_cells(start_row=status_title_row + 1, start_column=1, end_row=status_title_row + 1, end_column=6)
    ws.cell(status_title_row + 1, 1).value = (
        "STATUSのラベルと色、ACTION HISTORYの色は、本評価パネルの表示規約です。"
        "MetaDrive全体の共通色ではありません。"
        "セルの説明は自由に編集できます。"
    )
    ws.cell(status_title_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(status_title_row + 1, 1).font = Font(name="Yu Gothic", size=10, color=BLACK)
    ws.row_dimensions[status_title_row + 1].height = 30

    status_header_row = status_title_row + 3
    status_headers = ("表示", "対応フラグ", "説明", "表示色", "色コード")
    for column, value in enumerate(status_headers, start=1):
        ws.cell(status_header_row, column).value = value
    _style_header_row(ws, status_header_row, 1, 5)
    statuses = [
        ("RUNNING", "本評価の派生表示", "上記の判定条件が成立していない継続中の表示です。", "緑", GREEN),
        ("SUCCESS", "arrive_dest", "MetaDriveのarrive_destが真になった目的地到達です。", "緑", GREEN),
        ("OUT_OF_ROAD", "out_of_road", "out_of_road。設定により車線外だけでなく、連続線・歩道・route外なども含み得ます。", "赤", RED),
        ("CRASH_VEHICLE", "crash_vehicle", "crash_vehicle。他の車両との衝突フラグです。", "赤", RED),
        ("CRASH_OBJECT", "crash_object", "crash_object。障害物などの交通物体（traffic object）との衝突です。building・human・sidewalkは別flagです。", "赤", RED),
        ("CRASH", "crash", "crash。vehicle/object/building/human/sidewalk等をまとめた衝突の総合フラグです。個別flagがあればそちらを優先表示します。", "赤", RED),
        ("MAX_STEP", "truncated / max_step", "horizonによる時間上限の打切りです。実装上はtruncated=Trueまたはmax_stepで、このMetaDriveでは通常horizon timeout時にtruncated=Trueになります。", "黄", "F0BE46"),
        ("TERMINATED", "terminated（本評価の代替判定）", "上記の個別原因に当たらずterminated=Trueになった場合の本評価の代替判定です。公式のTerminationState名そのものではありません。", "赤", RED),
    ]
    status_data_start = status_header_row + 1
    for row_index, values in enumerate(statuses, start=status_data_start):
        for column, value in enumerate(values, start=1):
            ws.cell(row_index, column).value = value
        ws.row_dimensions[row_index].height = 48
    status_data_end = status_header_row + len(statuses)
    _style_body(ws, status_data_start, status_data_end, 1, 5)
    for row_index in range(status_data_start, status_data_end + 1):
        color_code = str(ws.cell(row_index, 5).value)
        ws.cell(row_index, 1).font = Font(name="Yu Gothic", size=10, bold=True, color=color_code)
        ws.cell(row_index, 4).fill = _fill(color_code)
        ws.cell(row_index, 4).font = Font(name="Yu Gothic", size=10, color=WHITE, bold=True)
        ws.cell(row_index, 5).font = Font(name="Consolas", size=10, color=BLACK)
    _add_table(
        ws,
        name="StatusLegend",
        ref=f"A{status_header_row}:E{status_data_end}",
    )

    action_title_row = status_data_end + 3
    ws.merge_cells(start_row=action_title_row, start_column=1, end_row=action_title_row, end_column=6)
    ws.cell(action_title_row, 1).value = "ACTION HISTORY の色"
    ws.cell(action_title_row, 1).font = Font(name="Yu Gothic", size=13, bold=True, color=CYAN)
    ws.cell(action_title_row, 1).fill = _fill(CYAN_SOFT)
    action_header_row = action_title_row + 1
    action_headers = ("帯", "値・向き", "色", "色コード", "説明")
    for column, value in enumerate(action_headers, start=1):
        ws.cell(action_header_row, column).value = value
    _style_header_row(ws, action_header_row, 1, 5)
    action_colors = [
        ("S", "positive / LEFT（左操舵）", "青", BLUE, "上段セル。本評価独自の色規約"),
        ("S", "negative / RIGHT（右操舵）", "オレンジ", "F5A541", "上段セル。本評価独自の色規約"),
        ("S", "zero", "グレー", HISTORY_GRAY, "上段セル。本評価独自の色規約"),
        ("T/B", "positive / THROTTLE（加速）", "緑", "37B969", "下段セル。本評価独自の色規約"),
        ("T/B", "negative / BRAKE（減速）", "赤", "E14B4B", "下段セル。本評価独自の色規約。reverse有効時は後退力になり得る"),
        ("T/B", "zero", "グレー", HISTORY_GRAY, "下段セル。本評価独自の色規約。0入力でも完全な無作動とは限らない"),
    ]
    action_data_start = action_header_row + 1
    for row_index, values in enumerate(action_colors, start=action_data_start):
        for column, value in enumerate(values, start=1):
            ws.cell(row_index, column).value = value
        ws.row_dimensions[row_index].height = 40
    action_data_end = action_header_row + len(action_colors)
    _style_body(ws, action_data_start, action_data_end, 1, 5)
    for row_index in range(action_data_start, action_data_end + 1):
        code = str(ws.cell(row_index, 4).value)
        ws.cell(row_index, 3).fill = _fill(code)
        ws.cell(row_index, 3).font = Font(name="Yu Gothic", size=10, bold=True, color=WHITE)
        ws.cell(row_index, 4).font = Font(name="Consolas", size=10, color=BLACK)
    _add_table(
        ws,
        name="ActionColorLegend",
        ref=f"A{action_header_row}:E{action_data_end}",
    )

    timing_title_row = action_data_end + 3
    ws.merge_cells(start_row=timing_title_row, start_column=1, end_row=timing_title_row, end_column=6)
    ws.cell(timing_title_row, 1).value = "メディア時間とシミュレータ時間"
    ws.cell(timing_title_row, 1).font = Font(name="Yu Gothic", size=13, bold=True, color=ORANGE)
    ws.cell(timing_title_row, 1).fill = _fill(ORANGE_SOFT)
    _write_merged_text(
        ws,
        f"A{timing_title_row + 1}:F{timing_title_row + 3}",
        "GIFとMP4は1制御ステップを1フレームとして出力します。各フレームの再生時間はCONTROLの逆数（1制御周期）と同じです。"
        "従ってN枚を再生し終えた累積時間はN×1制御周期で、N枚目のPOST-STEP tと一致します。"
        "生成・読み込みに要した実時間（wall clock）は含めません。",
        fill=ORANGE_SOFT,
    )
    for row in range(timing_title_row + 1, timing_title_row + 4):
        ws.row_dimensions[row].height = 25

    ws.freeze_panes = "A5"
    ws.print_area = f"A1:F{timing_title_row + 3}"


def _build_workbook(
    *,
    image: PILImage.Image,
    example: Mapping[str, Any],
) -> Workbook:
    workbook = Workbook()
    guide = workbook.active
    guide.title = "テレメトリ説明"

    _write_combined_guide(guide, image, example)
    _validate_workbook_tables(workbook)

    workbook.properties.title = "MetaDrive 評価テレメトリ表示ガイド"
    workbook.properties.subject = "frame_000030の右パネルと評価時間の読み方"
    workbook.properties.creator = "metadrive_rl"
    workbook.properties.description = "編集可能な日本語Excel資料。外部リンク・マクロ・外部ブック参照なし。"
    return workbook


def generate_guide(
    *,
    output_path: Path = DEFAULT_OUTPUT_PATH,
    frame_path: Path = DEFAULT_FRAME_PATH,
    steps_path: Path = DEFAULT_STEPS_PATH,
    evaluation_path: Path = DEFAULT_EVALUATION_PATH,
) -> Path:
    frame_path = _resolve_input(frame_path)
    steps_path = _resolve_input(steps_path)
    evaluation_path = _resolve_input(evaluation_path)
    output_path = _resolve_output(output_path)

    evaluation = _read_json(evaluation_path)
    rows = _read_step_rows(steps_path)
    example = _example_row(rows)
    config = _runtime_config(evaluation)
    history = _action_history(rows, example, config)
    image = _updated_frame(frame_path, example, history)

    workbook = _build_workbook(image=image, example=example)
    temporary_path = output_path.with_name(
        f"{output_path.stem}.tmp{output_path.suffix}"
    )
    temporary_path.unlink(missing_ok=True)
    try:
        workbook.save(temporary_path)
        temporary_path.replace(output_path)
    except BaseException:
        temporary_path.unlink(missing_ok=True)
        raise
    return output_path


def main() -> int:
    args = _parse_args()
    output_path = generate_guide(
        output_path=args.output,
        frame_path=args.frame,
        steps_path=args.steps,
        evaluation_path=args.evaluation,
    )
    print(f"Excel guide written: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
